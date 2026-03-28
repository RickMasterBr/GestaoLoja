"""
relatorios/excel_gerador.py — Geração de planilhas Excel para exportação.
Abre o arquivo diretamente no aplicativo padrão do Windows via os.startfile().
"""

import os
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Paleta de cores ────────────────────────────────────────────────────────────

HEADER_BG  = "1a3a5c"
HEADER_FG  = "FFFFFF"
SECTION_BG = "2d2d2d"
SECTION_FG = "FFFFFF"
ROW_ALT    = "F5F5F5"
TOTAL_BG   = "E8E8E8"
GREEN_FG   = "2d7a2d"
RED_FG     = "a02020"
ORANGE_FG  = "c47000"
GREY_FG    = "888888"

_PLAT_NOMES = {
    "iFood1": "iFood L1",
    "iFood2": "iFood L2",
    "99Food": "99Food",
    "Keeta":  "Keeta",
}


# ── Auxiliares ────────────────────────────────────────────────────────────────

def _r(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _borda_fina() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _borda_cab() -> Border:
    s = Side(style="thin", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)


# ── ExcelBuilder ───────────────────────────────────────────────────────────────

class ExcelBuilder:
    """Constrói planilhas Excel de forma incremental, linha a linha."""

    def __init__(self, titulo: str):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = titulo[:31]
        self.linha_atual = 1
        self._col_widths: dict = {}

    def adicionar_titulo(self, texto: str, subtitulo: str = None):
        cel = self.ws.cell(row=self.linha_atual, column=1, value=texto)
        cel.font = Font(name="Calibri", size=16, bold=True, color=HEADER_FG)
        cel.fill = PatternFill("solid", fgColor=HEADER_BG)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.merge_cells(f"A{self.linha_atual}:H{self.linha_atual}")
        self.ws.row_dimensions[self.linha_atual].height = 26
        self.linha_atual += 1

        if subtitulo:
            cel2 = self.ws.cell(row=self.linha_atual, column=1, value=subtitulo)
            cel2.font = Font(name="Calibri", size=11, italic=True, color="555555")
            cel2.alignment = Alignment(horizontal="center", vertical="center")
            self.ws.merge_cells(f"A{self.linha_atual}:H{self.linha_atual}")
            self.ws.row_dimensions[self.linha_atual].height = 18
            self.linha_atual += 1

        self.linha_atual += 1  # linha vazia

    def adicionar_secao(self, titulo: str):
        cel = self.ws.cell(row=self.linha_atual, column=1, value=titulo)
        cel.font = Font(name="Calibri", size=11, bold=True, color=SECTION_FG)
        cel.fill = PatternFill("solid", fgColor=SECTION_BG)
        cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        self.ws.merge_cells(f"A{self.linha_atual}:H{self.linha_atual}")
        self.ws.row_dimensions[self.linha_atual].height = 18
        self.linha_atual += 1

    def adicionar_cabecalho(self, colunas: list):
        brd = _borda_cab()
        for i, texto in enumerate(colunas, 1):
            cel = self.ws.cell(row=self.linha_atual, column=i, value=str(texto))
            cel.font = Font(name="Calibri", size=10, bold=True, color=HEADER_FG)
            cel.fill = PatternFill("solid", fgColor=HEADER_BG)
            cel.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cel.border = brd
            col_l = get_column_letter(i)
            self._col_widths[col_l] = max(
                self._col_widths.get(col_l, 10), min(len(str(texto)) + 4, 60)
            )
        self.linha_atual += 1

    def adicionar_linha(self, valores: list, alternada: bool = False,
                        negrito: bool = False, cor_texto: str = None,
                        is_total: bool = False, cores_por_celula: list = None):
        brd = _borda_fina()
        for i, valor in enumerate(valores, 1):
            cel = self.ws.cell(row=self.linha_atual, column=i, value=valor)

            # Cor desta célula: cores_por_celula sobrepõe cor_texto
            if cores_por_celula and (i - 1) < len(cores_por_celula) and cores_por_celula[i - 1]:
                cor_cel = cores_por_celula[i - 1]
            else:
                cor_cel = cor_texto

            font_kw: dict = {"name": "Calibri", "size": 10, "bold": negrito or is_total}
            if cor_cel:
                font_kw["color"] = cor_cel
            cel.font = Font(**font_kw)

            if is_total:
                cel.fill = PatternFill("solid", fgColor=TOTAL_BG)
            elif alternada:
                cel.fill = PatternFill("solid", fgColor=ROW_ALT)

            is_right = isinstance(valor, float) or (
                isinstance(valor, str) and valor.startswith("R$")
            )
            cel.alignment = Alignment(
                horizontal="right" if is_right else "left",
                vertical="center",
            )
            cel.border = brd

            col_l = get_column_letter(i)
            self._col_widths[col_l] = max(
                self._col_widths.get(col_l, 10), min(len(str(valor or "")) + 2, 60)
            )
        self.linha_atual += 1

    def adicionar_linha_vazia(self):
        self.linha_atual += 1

    def adicionar_nota(self, texto: str):
        cel = self.ws.cell(row=self.linha_atual, column=1, value=texto)
        cel.font = Font(name="Calibri", size=9, italic=True, color=GREY_FG)
        cel.alignment = Alignment(horizontal="left", vertical="center")
        self.ws.merge_cells(f"A{self.linha_atual}:H{self.linha_atual}")
        self.linha_atual += 1

    def auto_ajustar_colunas(self):
        for col_letter, width in self._col_widths.items():
            self.ws.column_dimensions[col_letter].width = max(10, min(width, 60))

    def salvar(self, nome_arquivo: str) -> str:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome = f"{nome_arquivo}_{ts}.xlsx"
        caminho = os.path.join(tempfile.gettempdir(), nome)
        self.auto_ajustar_colunas()
        self.wb.save(caminho)
        return caminho

    def abrir(self, caminho: str):
        os.startfile(caminho)


# ── Bloco de plataforma (compartilhado diário/período) ────────────────────────

def _bloco_plataforma_excel(b: ExcelBuilder, nome_plat: str, d: dict,
                             com_repasse: bool = False) -> None:
    b.adicionar_nota(_PLAT_NOMES.get(nome_plat, nome_plat))
    b.adicionar_cabecalho(["Descrição", "Valor"])
    cp  = d.get("comissao_pct", 0.0)
    tp  = d.get("tx_trans_pct", 0.0)
    spp = d.get("subsidio_pp",  0.0)
    lin = [
        [f"Pedidos: {d.get('qtd', 0)}  |  Bruto total", _r(d.get("bruto", 0.0))],
        ["Pago online (plataforma repassa)",              _r(d.get("bruto_online", 0.0))],
        [f"  (-) Comissão {cp:.1f}% s/ online",          _r(d.get("comissao_online", 0.0))],
        [f"  (-) Taxa transação {tp:.1f}% s/ online",    _r(d.get("tx_trans", 0.0))],
        [f"  (+) Subsídio R$ {spp:.2f}/ped",             _r(d.get("subsidio", 0.0))],
    ]
    if com_repasse:
        lin.append([f"Previsão de repasse: {d.get('dt_repasse', '—')}", ""])
    for i, row in enumerate(lin):
        cor = RED_FG if "(-)" in row[0] else None
        b.adicionar_linha(row, alternada=(i % 2 == 1), cor_texto=cor)
    b.adicionar_linha(
        ["Líquido Estimado", _r(d.get("liquido", 0.0))],
        is_total=True, cor_texto=GREEN_FG,
    )
    b.adicionar_linha_vazia()


# ══════════════════════════════════════════════════════════════════════════════
#  1. excel_relatorio_diario
# ══════════════════════════════════════════════════════════════════════════════

def excel_relatorio_diario(data_br: str, dados: dict) -> str:
    b = ExcelBuilder("Relatório Diário")
    b.adicionar_titulo(
        f"Relatório Diário — {data_br}",
        subtitulo=dados.get("nome_loja", "Gestão Loja"),
    )

    # Resumo por Canal
    b.adicionar_secao("Resumo por Canal")
    canais = dados.get("canais", [])
    if canais:
        b.adicionar_cabecalho(["Canal", "Qtd Pedidos", "Valor Total"])
        total_q = total_v = 0
        for i, r in enumerate(canais):
            qtd = r.get("qtd", 0)
            val = r.get("valor_liquido", 0.0)
            total_q += qtd
            total_v += val
            b.adicionar_linha(
                [r.get("canal_amigavel", r.get("canal", "")), str(qtd), _r(val)],
                alternada=(i % 2 == 1),
            )
        b.adicionar_linha(["TOTAL", str(total_q), _r(total_v)], is_total=True)
    else:
        b.adicionar_nota("Sem registros para esta data.")
    b.adicionar_linha_vazia()

    # Pagamentos
    b.adicionar_secao("Pagamentos")
    b.adicionar_nota("VA/VR = Benefício · Voucher/Cortesia excluídos")
    pagamentos = dados.get("pagamentos", [])
    if pagamentos:
        b.adicionar_cabecalho(["Método", "Tipo", "Valor Total"])
        total_p = 0.0
        for i, r in enumerate(pagamentos):
            val = r.get("total", 0.0)
            total_p += val
            b.adicionar_linha(
                [r.get("metodo", ""), r.get("tipo", ""), _r(val)],
                alternada=(i % 2 == 1),
            )
        b.adicionar_linha(["TOTAL", "", _r(total_p)], is_total=True)
    else:
        b.adicionar_nota("Sem registros para esta data.")
    b.adicionar_linha_vazia()

    # Detalhamento Plataformas
    b.adicionar_secao("Detalhamento Plataformas")
    plataformas = dados.get("plataformas", {})
    alguma = False
    for np in ["iFood1", "iFood2", "99Food", "Keeta"]:
        d = plataformas.get(np, {})
        if not d.get("qtd", 0):
            continue
        alguma = True
        _bloco_plataforma_excel(b, np, d, com_repasse=False)
    if not alguma:
        b.adicionar_nota("Sem registros para esta data.")
    b.adicionar_linha_vazia()

    # Entregadores
    b.adicionar_secao("Entregadores")
    entregadores = dados.get("entregadores", [])
    if entregadores:
        b.adicionar_cabecalho([
            "Nome", "Entregas", "Soma Taxas", "Diária",
            "Extras", "Vales", "Total a Pagar",
        ])
        for i, r in enumerate(entregadores):
            b.adicionar_linha([
                r.get("nome", ""),
                str(r.get("total_entregas", 0)),
                _r(r.get("soma_taxas", 0.0)),
                _r(r.get("diaria", 0.0)),
                _r(r.get("corridas_extras", 0.0)),
                _r(r.get("vales", 0.0)),
                _r(r.get("total_liquido", 0.0)),
            ], alternada=(i % 2 == 1))
    else:
        b.adicionar_nota("Sem registros para esta data.")
    b.adicionar_linha_vazia()

    # Fechamento de Caixa
    b.adicionar_secao("Fechamento de Caixa")
    cx  = dados.get("caixa", {})
    dif = cx.get("diferenca", 0.0)
    b.adicionar_cabecalho(["Descrição", "Valor"])
    for i, row in enumerate([
        ["Troco Inicial",        _r(cx.get("troco_inicial", 0.0))],
        ["Entradas Espécie",     _r(cx.get("total_especie_entradas", 0.0))],
        ["Saídas Espécie",       _r(cx.get("total_especie_saidas", 0.0))],
        ["Saldo Teórico",        _r(cx.get("saldo_teorico", 0.0))],
        ["Saldo Real (gaveta)",  _r(cx.get("saldo_gaveta_real", 0.0))],
    ]):
        b.adicionar_linha(row, alternada=(i % 2 == 1))
    cor_dif = GREEN_FG if dif == 0 else (RED_FG if dif < 0 else ORANGE_FG)
    b.adicionar_linha(["Diferença", _r(dif)], is_total=True, cor_texto=cor_dif)
    b.adicionar_linha_vazia()

    # Movimentações do Dia
    b.adicionar_secao("Movimentações do Dia")
    extras = dados.get("extras", [])
    if extras:
        b.adicionar_cabecalho(["Pessoa", "Categoria", "Fluxo", "Método", "Valor", "Obs"])
        for i, r in enumerate(extras):
            fluxo = r.get("fluxo", "")
            cor = GREEN_FG if fluxo == "ENTRADA" else (RED_FG if fluxo == "SAIDA" else None)
            b.adicionar_linha([
                r.get("nome_pessoa", "—"),
                r.get("categoria", ""),
                fluxo,
                r.get("metodo", "—"),
                _r(r.get("valor", 0.0)),
                r.get("obs", ""),
            ], alternada=(i % 2 == 1), cor_texto=cor)
    else:
        b.adicionar_nota("Sem registros para esta data.")

    return b.salvar("relatorio_diario")


# ══════════════════════════════════════════════════════════════════════════════
#  2. excel_relatorio_periodo
# ══════════════════════════════════════════════════════════════════════════════

def excel_relatorio_periodo(ini_br: str, fim_br: str, dados: dict) -> str:
    b = ExcelBuilder("Relatório Período")
    b.adicionar_titulo(
        f"Relatório de Período — {ini_br} a {fim_br}",
        subtitulo=dados.get("nome_loja", "Gestão Loja"),
    )

    # Resumo Geral
    b.adicionar_secao("Resumo Geral")
    rg = dados.get("resumo_geral", {})
    b.adicionar_cabecalho(["Métrica", "Valor"])
    for i, row in enumerate([
        ["Total de Pedidos", str(rg.get("total_pedidos", 0))],
        ["Valor Bruto",      _r(rg.get("valor_bruto", 0.0))],
        ["Faturamento Real", _r(rg.get("fat_real", 0.0))],
        ["Total Cortesias",  _r(rg.get("total_cortesias", 0.0))],
        ["Taxas de Entrega", _r(rg.get("total_taxas", 0.0))],
    ]):
        b.adicionar_linha(row, alternada=(i % 2 == 1))
    b.adicionar_linha_vazia()

    # Resumo por Canal
    b.adicionar_secao("Resumo por Canal")
    canais = dados.get("canais", [])
    if canais:
        b.adicionar_cabecalho(["Canal", "Qtd Pedidos", "Valor Total"])
        total_q = total_v = 0
        for i, r in enumerate(canais):
            qtd = r.get("qtd", 0)
            val = r.get("valor_total", r.get("valor_liquido", 0.0))
            total_q += qtd
            total_v += val
            b.adicionar_linha(
                [r.get("canal_amigavel", r.get("canal", "")), str(qtd), _r(val)],
                alternada=(i % 2 == 1),
            )
        b.adicionar_linha(["TOTAL", str(total_q), _r(total_v)], is_total=True)
    else:
        b.adicionar_nota("Sem registros para este período.")
    b.adicionar_linha_vazia()

    # Resumo por Pagamento
    b.adicionar_secao("Resumo por Pagamento")
    b.adicionar_nota("VA/VR = Benefício · Voucher/Cortesia excluídos")
    pagamentos = dados.get("pagamentos", [])
    if pagamentos:
        b.adicionar_cabecalho(["Método", "Tipo", "Valor Total"])
        total_p = 0.0
        for i, r in enumerate(pagamentos):
            val = r.get("total", 0.0)
            total_p += val
            b.adicionar_linha(
                [r.get("metodo", ""), r.get("tipo", ""), _r(val)],
                alternada=(i % 2 == 1),
            )
        b.adicionar_linha(["TOTAL", "", _r(total_p)], is_total=True)
    else:
        b.adicionar_nota("Sem registros para este período.")
    b.adicionar_linha_vazia()

    # Detalhamento Plataformas
    b.adicionar_secao("Detalhamento Plataformas")
    plataformas = dados.get("plataformas", {})
    alguma = False
    for np in ["iFood1", "iFood2", "99Food", "Keeta"]:
        d = plataformas.get(np, {})
        if not d.get("qtd", 0):
            continue
        alguma = True
        _bloco_plataforma_excel(b, np, d, com_repasse=True)
    if not alguma:
        b.adicionar_nota("Sem registros para este período.")
    b.adicionar_linha_vazia()

    # Entregadores no Período
    b.adicionar_secao("Entregadores no Período")
    entregadores = dados.get("entregadores", [])
    if entregadores:
        b.adicionar_cabecalho([
            "Nome", "Entregas", "Soma Taxas", "Diárias",
            "Extras", "Vales", "Total a Pagar",
        ])
        for i, r in enumerate(entregadores):
            b.adicionar_linha([
                r.get("nome", ""),
                str(r.get("total_entregas", 0)),
                _r(r.get("soma_taxas", 0.0)),
                _r(r.get("total_diarias", r.get("diaria", 0.0))),
                _r(r.get("corridas_extras", 0.0)),
                _r(r.get("vales", 0.0)),
                _r(r.get("total_liquido", 0.0)),
            ], alternada=(i % 2 == 1))
    else:
        b.adicionar_nota("Sem registros para este período.")
    b.adicionar_linha_vazia()

    # Funcionários no Período
    b.adicionar_secao("Funcionários no Período")
    funcionarios = dados.get("funcionarios", [])
    if funcionarios:
        b.adicionar_cabecalho(["Nome", "Cargo", "Tipo Salário", "Total Estimado"])
        for i, r in enumerate(funcionarios):
            b.adicionar_linha([
                r.get("nome", ""),
                r.get("cargo", ""),
                r.get("tipo_salario", ""),
                _r(r.get("total_estimado", 0.0)),
            ], alternada=(i % 2 == 1))
    else:
        b.adicionar_nota("Sem registros para este período.")
    b.adicionar_linha_vazia()

    # Projeção de Repasses
    b.adicionar_secao("Projeção de Repasses")
    b.adicionar_cabecalho(["Plataforma", "Líquido Estimado", "Data Prevista"])
    for i, np in enumerate(["iFood1", "iFood2", "99Food", "Keeta"]):
        d = plataformas.get(np, {})
        b.adicionar_linha([
            _PLAT_NOMES.get(np, np),
            _r(d.get("liquido", 0.0)),
            d.get("dt_repasse", "—"),
        ], alternada=(i % 2 == 1))

    return b.salvar("relatorio_periodo")


# ══════════════════════════════════════════════════════════════════════════════
#  3. excel_fluxo_caixa
# ══════════════════════════════════════════════════════════════════════════════

def excel_fluxo_caixa(titulo: str, ini_br: str, fim_br: str,
                       lancamentos: list) -> str:
    b = ExcelBuilder("Fluxo de Caixa")
    subtitulo = ini_br if ini_br == fim_br else f"{ini_br} a {fim_br}"
    b.adicionar_titulo(f"Fluxo de Caixa — {titulo}", subtitulo=subtitulo)

    total_e = sum((r.get("entrada") or 0.0) for r in lancamentos)
    total_s = sum((r.get("saida")   or 0.0) for r in lancamentos)
    saldo_f = total_e - total_s

    b.adicionar_secao("Resumo")
    b.adicionar_cabecalho(["Métrica", "Valor"])
    b.adicionar_linha(["Total Entradas", _r(total_e)], cor_texto=GREEN_FG)
    b.adicionar_linha(["Total Saídas",   _r(total_s)], alternada=True, cor_texto=RED_FG)
    b.adicionar_linha(
        ["Saldo Final", _r(saldo_f)],
        is_total=True,
        cor_texto=GREEN_FG if saldo_f >= 0 else RED_FG,
    )
    b.adicionar_linha_vazia()

    b.adicionar_secao("Lançamentos")
    b.adicionar_cabecalho([
        "Data", "Hora", "Tipo", "Descrição", "Método",
        "Entrada", "Saída", "Saldo",
    ])
    saldo = 0.0
    for i, r in enumerate(lancamentos):
        entrada = r.get("entrada") or 0.0
        saida   = r.get("saida")   or 0.0
        saldo  += entrada - saida
        cor_row  = GREEN_FG if entrada > 0 else (RED_FG if saida > 0 else None)
        cor_saldo = GREEN_FG if saldo >= 0 else RED_FG
        b.adicionar_linha(
            [
                r.get("data", ""),
                r.get("hora", "") or "",
                r.get("tipo", ""),
                r.get("descricao", "") or "",
                r.get("metodo", "") or "",
                _r(entrada) if entrada else "",
                _r(saida)   if saida   else "",
                _r(saldo),
            ],
            alternada=(i % 2 == 1),
            cor_texto=cor_row,
            cores_por_celula=[None, None, None, None, None, None, None, cor_saldo],
        )

    return b.salvar("fluxo_caixa")


# ══════════════════════════════════════════════════════════════════════════════
#  4. excel_divergencias
# ══════════════════════════════════════════════════════════════════════════════

def excel_divergencias(ini_br: str, fim_br: str, registros: list) -> str:
    b = ExcelBuilder("Divergências")
    b.adicionar_titulo(f"Histórico de Divergências — {ini_br} a {fim_br}")

    total = len(registros)
    n_div = sum(1 for r in registros if abs(r.get("diferenca") or 0.0) > 0.001)

    b.adicionar_secao("Resumo")
    b.adicionar_cabecalho(["Métrica", "Valor"])
    b.adicionar_linha(["Total de Dias no Período", str(total)])
    b.adicionar_linha(["Dias com Divergência", str(n_div)], alternada=True,
                      cor_texto=RED_FG if n_div > 0 else None)
    b.adicionar_linha_vazia()

    b.adicionar_secao("Fechamentos")
    b.adicionar_cabecalho([
        "Data", "Saldo Teórico", "Saldo Real", "Diferença", "Observação",
    ])
    for i, r in enumerate(registros):
        dif = r.get("diferenca") or 0.0
        cor_dif = GREEN_FG if abs(dif) <= 0.001 else (RED_FG if dif < 0 else ORANGE_FG)
        b.adicionar_linha(
            [
                r.get("data", ""),
                _r(r.get("saldo_teorico") or 0.0),
                _r(r.get("saldo_gaveta_real") or 0.0),
                _r(dif),
                r.get("obs_fechamento") or "",
            ],
            alternada=(i % 2 == 1),
            cores_por_celula=[None, None, None, cor_dif, None],
        )

    return b.salvar("divergencias_fechamento")


# ══════════════════════════════════════════════════════════════════════════════
#  5. excel_holerite
# ══════════════════════════════════════════════════════════════════════════════

def excel_holerite(nome: str, mes_ano: str, dados: dict) -> str:
    b = ExcelBuilder("Holerite")
    b.adicionar_titulo(f"Holerite — {nome} — {mes_ano}")

    # Resumo
    b.adicionar_secao("Resumo")
    resumo = dados.get("resumo", [])
    if resumo:
        b.adicionar_cabecalho(["Descrição", "Valor"])
        for i, r in enumerate(resumo):
            tipo   = r.get("tipo", "")
            is_tot = tipo == "total"
            cor    = GREEN_FG if is_tot else (RED_FG if tipo == "desconto" else None)
            b.adicionar_linha(
                [r.get("descricao", ""), _r(r.get("valor", 0.0))],
                alternada=(i % 2 == 1) and not is_tot,
                cor_texto=cor,
                is_total=is_tot,
                negrito=is_tot,
            )
    b.adicionar_linha_vazia()

    # Vales
    b.adicionar_secao("Detalhamento Vales")
    vales = dados.get("vales", [])
    if vales:
        b.adicionar_cabecalho(["Data", "Valor", "Observação"])
        for i, r in enumerate(vales):
            b.adicionar_linha(
                [r.get("data", ""), _r(r.get("valor", 0.0)), r.get("obs", "")],
                alternada=(i % 2 == 1),
            )
    else:
        b.adicionar_nota("Nenhum vale registrado.")
    b.adicionar_linha_vazia()

    # Consumos
    b.adicionar_secao("Detalhamento Consumos")
    consumos = dados.get("consumos", [])
    if consumos:
        b.adicionar_cabecalho(["Data", "Valor Original", "Desconto 80%", "Observação"])
        for i, r in enumerate(consumos):
            b.adicionar_linha([
                r.get("data", ""),
                _r(r.get("valor_original", 0.0)),
                _r(r.get("desconto_80", 0.0)),
                r.get("obs", ""),
            ], alternada=(i % 2 == 1))
    else:
        b.adicionar_nota("Nenhum consumo registrado.")
    b.adicionar_linha_vazia()

    # Ocorrências
    b.adicionar_secao("Ocorrências de Escala")
    ocorrencias = dados.get("ocorrencias", [])
    if ocorrencias:
        b.adicionar_cabecalho(["Data", "Tipo", "Impacto"])
        for i, r in enumerate(ocorrencias):
            b.adicionar_linha(
                [r.get("data", ""), r.get("tipo", ""), r.get("impacto", "")],
                alternada=(i % 2 == 1),
            )
    else:
        b.adicionar_nota("Nenhuma ocorrência registrada.")
    b.adicionar_linha_vazia()

    # Controle de Ponto
    ponto = dados.get("ponto", [])
    if ponto:
        b.adicionar_secao("Controle de Ponto")
        b.adicionar_cabecalho([
            "Data", "Entrada", "Saída", "H.Brutas",
            "H.Líquidas", "Extras/Faltantes",
        ])
        for i, r in enumerate(ponto):
            ev  = str(r.get("extras_faltantes", ""))
            cor = GREEN_FG if ev.startswith("+") else (RED_FG if ev.startswith("-") else None)
            b.adicionar_linha([
                r.get("data", ""),
                r.get("entrada", ""),
                r.get("saida", ""),
                r.get("horas_brutas", ""),
                r.get("horas_liquidas", ""),
                ev,
            ], alternada=(i % 2 == 1),
               cores_por_celula=[None, None, None, None, None, cor])

    return b.salvar("holerite")


# ══════════════════════════════════════════════════════════════════════════════
#  6. excel_entregadores
# ══════════════════════════════════════════════════════════════════════════════

def excel_entregadores(data_br: str, dados: dict) -> str:
    b = ExcelBuilder("Entregadores")
    b.adicionar_titulo(f"Painel Entregadores — {data_br}")

    COLUNAS = [
        "Nome", "Entregas", "Soma Taxas", "Diária",
        "Corridas Extra", "Vales", "Total a Pagar",
    ]

    def _bloco(titulo_secao: str, lista: list, msg_vazio: str):
        b.adicionar_secao(titulo_secao)
        if lista:
            b.adicionar_cabecalho(COLUNAS)
            for i, r in enumerate(lista):
                b.adicionar_linha([
                    r.get("nome", ""),
                    str(r.get("entregas", r.get("total_entregas", 0))),
                    _r(r.get("soma_taxas", 0.0)),
                    _r(r.get("diaria", 0.0)),
                    _r(r.get("corridas_extras", r.get("corridas_extra", 0.0))),
                    _r(r.get("vales", 0.0)),
                    _r(r.get("total_a_pagar", r.get("total_liquido", 0.0))),
                ], alternada=(i % 2 == 1))
        else:
            b.adicionar_nota(msg_vazio)
        b.adicionar_linha_vazia()

    _bloco("Resumo do Dia",        dados.get("dia", []),    "Sem registros para esta data.")
    _bloco("Acumulado da Semana",  dados.get("semana", []), "Sem dados acumulados para a semana.")

    return b.salvar("entregadores")


# ══════════════════════════════════════════════════════════════════════════════
#  7. excel_estoque_movimentacoes
# ══════════════════════════════════════════════════════════════════════════════

def excel_estoque_movimentacoes(ini_br: str, fim_br: str,
                                 movimentacoes: list, resumo: dict) -> str:
    b = ExcelBuilder("Estoque Movimentações")
    b.adicionar_titulo(f"Estoque — Movimentações — {ini_br} a {fim_br}")

    b.adicionar_secao("Resumo")
    b.adicionar_cabecalho(["Descrição", "Qtd", "Valor Total"])
    b.adicionar_linha([
        "Total Entradas",
        str(resumo.get("total_entrada_qtd", 0)),
        _r(resumo.get("total_entrada_valor", 0.0)),
    ], cor_texto=GREEN_FG)
    b.adicionar_linha([
        "Total Saídas",
        str(resumo.get("total_saida_qtd", 0)),
        _r(resumo.get("total_saida_valor", 0.0)),
    ], alternada=True, cor_texto=RED_FG)
    b.adicionar_linha_vazia()

    b.adicionar_secao("Movimentações")
    if movimentacoes:
        b.adicionar_cabecalho([
            "Data", "Hora", "Produto", "Categoria", "Tipo",
            "Qtd", "Preço Unit.", "Valor Total", "Motivo", "Obs",
        ])
        for i, r in enumerate(movimentacoes):
            tipo = (r.get("tipo") or "").upper()
            cor  = GREEN_FG if tipo == "ENTRADA" else (RED_FG if tipo == "SAIDA" else None)
            b.adicionar_linha([
                r.get("data", ""),
                r.get("hora", "") or "",
                r.get("produto", r.get("nome_produto", "")),
                r.get("categoria", r.get("nome_categoria", "")),
                tipo,
                str(r.get("quantidade", 0)),
                _r(r.get("preco_unit", r.get("preco_unitario", 0.0))),
                _r(r.get("valor_total", 0.0)),
                r.get("motivo", "") or "",
                r.get("obs", r.get("observacao", "")) or "",
            ], alternada=(i % 2 == 1), cor_texto=cor)
    else:
        b.adicionar_nota("Sem movimentações para este período.")

    return b.salvar("estoque_movimentacoes")
